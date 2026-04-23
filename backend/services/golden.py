"""Golden data ingestion — reads analyst-created Excel files into normalized rows.

The golden data Excel follows a standard DD2 template:
  Row 1: Area headers (DD2 Information Area, File Information Area, etc.)
  Row 2: Required/Optional indicators
  Row 3: Column names (field names)
  Row 4+: Data rows

Column names vary slightly from our schema (e.g., "Quanity" vs "quantity").
This module handles the mapping + normalization.
"""

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import yaml

from backend.settings import settings

logger = logging.getLogger(__name__)


# ── Column name mapping: golden Excel header → our schema field name ──
# Built from reading real golden data Excel files (DD2 template format).
# Keys are normalized (lowercase, stripped). This handles minor spelling variations.

_COLUMN_MAP: dict[str, str] = {
    "status": "status",
    "inventory creation questions, concerns, or notes.": "notes",
    "*contract info received": "contract_info_received",
    "invoice file name": "invoice_file_name",
    "files used for inventory": "files_used",
    "billing name": "billing_name",
    "service address 1": "service_address_1",
    "service address 2": "service_address_2",
    "city": "city",
    "state": "state",
    "zip": "zip",
    "country": "country",
    "carrier": "carrier_name",
    "master account": "master_account",
    "carrier account number": "carrier_account_number",
    "sub-account number": "sub_account_number_1",
    "sub-account number 2": "sub_account_number_2",
    "btn": "btn",
    "phone number": "phone_number",
    "carrier circuit number": "carrier_circuit_number",
    "additional circuit ids": "additional_circuit_ids",
    "service type": "service_type",
    "service type 2": "service_type_2",
    "usoc": "usoc",
    "service or component": "service_or_component",
    "component or feature name": "component_or_feature_name",
    "monthly recurring cost": "monthly_recurring_cost",
    "quanity": "quantity",          # Note: golden data has typo "Quanity"
    "quantity": "quantity",
    "cost per unit": "cost_per_unit",
    "currency": "currency",
    "conversion rate": "conversion_rate",
    "monthly recurring cost per currency": "mrc_per_currency",
    "charge type": "charge_type",
    "# calls": "num_calls",
    "ld minutes": "ld_minutes",
    "ld cost": "ld_cost",
    "rate": "rate",
    "ld flat rate": "ld_flat_rate",
    "port speed": "port_speed",
    "access speed": "access_speed",
    "upload speed": "upload_speed",
    "poin to number": "point_to_number",     # Golden data typo
    "point to number": "point_to_number",
    "z location name if one given by carrier": "z_location_name",
    "z location name": "z_location_name",
    "z address 1": "z_address_1",
    "z address 2": "z_address_2",
    "z city": "z_city",
    "z state": "z_state",
    "z zip code": "z_zip",
    "z zip": "z_zip",
    "z country": "z_country",
    "*contract - term months": "contract_term_months",
    "contract - term months": "contract_term_months",
    "*contract - begin date": "contract_begin_date",
    "contract - begin date": "contract_begin_date",
    "*contract - expiration date": "contract_expiration_date",
    "contract - expiration date": "contract_expiration_date",
    "billing per contract": "billing_per_contract",
    "*currently month-to-month": "currently_month_to_month",
    "currently month-to-month": "currently_month_to_month",
    "month to month or less than a year remaining": "mtm_or_less_than_year",
    "contract file name": "contract_file_name",
    "contract number": "contract_number",
    "2nd contract number": "contract_number_2",
    "*auto renew": "auto_renew",
    "auto renew": "auto_renew",
    "auto renewal notes and removal requirements": "auto_renewal_notes",
    "auto renewal notes": "auto_renewal_notes",
}


def _normalize_header(h: str) -> str:
    """Normalize a column header for lookup: lowercase, strip, collapse whitespace."""
    if not h:
        return ""
    return re.sub(r'\s+', ' ', str(h).strip().lower())


def _normalize_cell_value(field_name: str, value) -> str | float | int | None:
    """Normalize a cell value based on target field type.

    Handles Excel artifacts: numeric phones with .0, datetime dates, etc.
    Returns the normalized value as a string (for consistency with our comparison).
    """
    if value is None:
        return None

    # Dates: Excel stores as datetime objects
    if field_name in ("contract_begin_date", "contract_expiration_date"):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.isoformat()
        s = str(value).strip()
        if not s:
            return None
        # Try parsing common formats
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s

    # Numeric fields stored as floats in Excel
    if field_name in ("monthly_recurring_cost", "cost_per_unit", "mrc_per_currency",
                       "ld_cost", "rate", "ld_flat_rate", "ld_minutes",
                       "conversion_rate"):
        if isinstance(value, (int, float)):
            return value
        s = str(value).strip().replace("$", "").replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return s

    if field_name in ("quantity", "num_calls", "contract_term_months"):
        if isinstance(value, (int, float)):
            return int(value)
        s = str(value).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return s

    # Phone numbers: Excel may store as numeric (e.g., 5551234567.0)
    if field_name in ("phone_number", "btn", "point_to_number"):
        s = str(value)
        # Strip .0 from Excel numeric representation
        if s.endswith(".0") and s[:-2].replace("-", "").replace(" ", "").isdigit():
            s = s[:-2]
        return s.strip() if s.strip() else None

    # Everything else: string
    s = str(value).strip()
    # Strip .0 from numeric-looking strings (account numbers stored as floats)
    if s.endswith(".0") and s[:-2].replace("-", "").replace(" ", "").isdigit():
        s = s[:-2]
    return s if s else None


def load_golden_excel(
    file_path: str,
    carriers: set[str] | None = None,
    sheet_name: str = "Baseline",
) -> list[dict]:
    """Load golden data from Excel file, return list of {schema_field: normalized_value}.

    Args:
        file_path: path to the golden Excel file
        carriers: optional filter — only return rows for these carriers (case-insensitive)
        sheet_name: which sheet to read (default "Baseline")

    Returns:
        List of dicts where keys are our 60-field schema names and values are normalized.
    """
    logger.info(f"Loading golden data from {file_path} (sheet={sheet_name})")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 4:
        raise ValueError(f"Expected at least 4 rows (2 header + 1 column names + data), got {len(all_rows)}")

    # Row 3 (index 2) has column names
    raw_headers = all_rows[2]
    col_map: list[tuple[int, str]] = []  # (col_index, schema_field_name)

    for i, header in enumerate(raw_headers):
        norm_h = _normalize_header(header)
        if norm_h in _COLUMN_MAP:
            col_map.append((i, _COLUMN_MAP[norm_h]))

    mapped_fields = [f for _, f in col_map]
    logger.info(f"Mapped {len(col_map)}/{len([h for h in raw_headers if h])} columns to schema fields")

    # Check for unmapped columns
    for i, header in enumerate(raw_headers):
        if header and _normalize_header(header) not in _COLUMN_MAP:
            logger.debug(f"Unmapped golden column {i}: {header!r}")

    # Parse data rows
    golden_rows = []
    for row in all_rows[3:]:
        record = {}
        for col_idx, field_name in col_map:
            if col_idx < len(row):
                val = _normalize_cell_value(field_name, row[col_idx])
                if val is not None:
                    record[field_name] = val

        # Skip empty rows
        if not record or not any(v for v in record.values()):
            continue

        # Filter by carrier (alias-aware: "Spectrum" also matches "Charter Communications")
        if carriers:
            row_carrier = str(record.get("carrier_name", "")).strip().lower()
            matched = False
            for c in carriers:
                # Direct substring match
                if c.lower() in row_carrier or row_carrier in c.lower():
                    matched = True
                    break
                # Alias-based match: check if the carrier config's aliases cover the golden value
                try:
                    from backend.config_loader import get_config_store
                    store = get_config_store()
                    carrier_config = store.get_carrier(c.lower())
                    if carrier_config:
                        if any(alias.lower() in row_carrier or row_carrier in alias.lower()
                               for alias in carrier_config.aliases):
                            matched = True
                            break
                except Exception:
                    pass
            if not matched:
                continue

        golden_rows.append(record)

    logger.info(f"Loaded {len(golden_rows)} golden rows"
                + (f" for carriers {carriers}" if carriers else ""))

    return golden_rows


def load_eval_config() -> dict:
    """Load eval configuration from configs/processing/eval_config.yaml."""
    config_path = Path(settings.configs_dir) / "processing" / "eval_config.yaml"
    if not config_path.exists():
        logger.warning(f"Eval config not found at {config_path}, using defaults")
        return {
            "default_field_extractability": {"analyst_judgment": [], "derived": []},
            "accuracy_targets": {"structured": 0.95, "semi_structured": 0.85, "fuzzy": 0.75, "contract": 0.70},
        }

    with open(config_path) as f:
        return yaml.safe_load(f) or {}


def classify_field_extractability(carrier: str | None = None) -> dict[str, str]:
    """Return {field_name: "extractable"|"analyst_judgment"|"derived"} for all 60 fields.

    Loads default from eval_config.yaml, then merges carrier-specific overrides
    from configs/carriers/{carrier}/eval.yaml if it exists.
    """
    from backend.models.schemas import FIELD_CATEGORIES

    config = load_eval_config()
    extractability = config.get("default_field_extractability", {})

    analyst_fields = set(extractability.get("analyst_judgment", []))
    derived_fields = set(extractability.get("derived", []))

    # Load carrier override if available
    if carrier:
        carrier_eval_path = Path(settings.configs_dir) / "carriers" / carrier / "eval.yaml"
        if carrier_eval_path.exists():
            with open(carrier_eval_path) as f:
                carrier_config = yaml.safe_load(f) or {}
            carrier_ext = carrier_config.get("field_extractability", {})
            analyst_fields |= set(carrier_ext.get("analyst_judgment", []))
            derived_fields |= set(carrier_ext.get("derived", []))
            # Allow carrier to override defaults (mark something as extractable)
            for field in carrier_ext.get("extractable", []):
                analyst_fields.discard(field)
                derived_fields.discard(field)

    # Build full classification
    result = {}
    for field_name in FIELD_CATEGORIES:
        if field_name in analyst_fields:
            result[field_name] = "analyst_judgment"
        elif field_name in derived_fields:
            result[field_name] = "derived"
        else:
            result[field_name] = "extractable"

    return result
