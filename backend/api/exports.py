"""Export/Import API — Excel download and corrected Excel upload.

The Excel export is driven by `configs/processing/export_template.yaml` so
it always matches the customer's master inventory template:
  Sheet 1 "Extracted Data" — the rows in customer column order, with a
                              "Required level" annotation row above headers
  Sheet 2 "Checklist"      — 30 QA items (Agent / QA Yes-No columns blank
                              for the analyst to fill in)
  Sheet 3 "Column Explanations" — verbatim column definitions
  Sheet 4 "Mandatory Fields"    — requirement matrix grouped by area
"""

import io
import logging
from functools import lru_cache
from pathlib import Path

import yaml
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.database import get_db
from backend.models.orm import ExtractedRow, ExtractionRun, Correction
from backend.settings import settings

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/exports", tags=["exports"])


@lru_cache(maxsize=1)
def _export_template() -> dict:
    """Load + cache the customer's export template config."""
    path = Path(settings.configs_dir) / "processing" / "export_template.yaml"
    if not path.exists():
        logger.warning("export_template.yaml missing — falling back to legacy column list")
        return {}
    return yaml.safe_load(path.read_text()) or {}


# Legacy column list — used as a fallback if the YAML is unavailable, and
# kept aliased to _EXCEL_FIELDS so /corrections/import (which expects this
# exact order) continues to work even after we switch to the YAML for export.
_EXCEL_FIELDS = [
    ("status", "Status"),
    ("notes", "Notes"),
    ("contract_info_received", "Contract Info Received"),
    ("invoice_file_name", "Invoice File Name"),
    ("files_used", "Files Used"),
    ("billing_name", "Billing Name"),
    ("service_address_1", "Service Address 1"),
    ("service_address_2", "Service Address 2"),
    ("city", "City"),
    ("state", "State"),
    ("zip", "Zip"),
    ("country", "Country"),
    ("carrier_name", "Carrier"),
    ("master_account", "Master Account"),
    ("carrier_account_number", "Carrier Account Number"),
    ("sub_account_number_1", "Sub-Account Number"),
    ("sub_account_number_2", "Sub-Account Number 2"),
    ("btn", "BTN"),
    ("phone_number", "Phone Number"),
    ("carrier_circuit_number", "Carrier Circuit Number"),
    ("additional_circuit_ids", "Additional Circuit IDs"),
    ("service_type", "Service Type"),
    ("service_type_2", "Service Type 2"),
    ("usoc", "USOC"),
    ("service_or_component", "Service or Component"),
    ("component_or_feature_name", "Component or Feature Name"),
    ("monthly_recurring_cost", "Monthly Recurring Cost"),
    ("quantity", "Quantity"),
    ("cost_per_unit", "Cost Per Unit"),
    ("currency", "Currency"),
    ("conversion_rate", "Conversion Rate"),
    ("mrc_per_currency", "MRC per Currency"),
    ("charge_type", "Charge Type"),
    ("num_calls", "# Calls"),
    ("ld_minutes", "LD Minutes"),
    ("ld_cost", "LD Cost"),
    ("rate", "Rate"),
    ("ld_flat_rate", "LD Flat Rate"),
    ("point_to_number", "Point to Number"),
    ("port_speed", "Port Speed"),
    ("access_speed", "Access Speed"),
    ("upload_speed", "Upload Speed"),
    ("z_location_name", "Z Location Name"),
    ("z_address_1", "Z Address 1"),
    ("z_address_2", "Z Address 2"),
    ("z_city", "Z City"),
    ("z_state", "Z State"),
    ("z_zip", "Z Zip"),
    ("z_country", "Z Country"),
    ("contract_term_months", "Contract Term Months"),
    ("contract_begin_date", "Contract Begin Date"),
    ("contract_expiration_date", "Contract Expiration Date"),
    ("billing_per_contract", "Billing Per Contract"),
    ("currently_month_to_month", "Currently Month-to-Month"),
    ("mtm_or_less_than_year", "MTM or Less Than Year"),
    ("contract_file_name", "Contract File Name"),
    ("contract_number", "Contract Number"),
    ("contract_number_2", "2nd Contract Number"),
    ("auto_renew", "Auto Renew"),
    ("auto_renewal_notes", "Auto Renewal Notes"),
]


@router.get("/{upload_id}/excel")
async def export_excel(upload_id: str, db: AsyncSession = Depends(get_db)):
    """Download extracted rows as a multi-sheet Excel matching the customer's
    master inventory template (configs/processing/export_template.yaml).

    Sheets:
      1. Extracted Data        — rows in customer column order + requirement
                                  level annotation row + confidence color coding
      2. Checklist             — 30 QA items, Agent/QA Yes-No columns
      3. Column Explanations   — verbatim column definitions
      4. Mandatory Fields      — requirement matrix grouped by area
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        # Find extraction run(s) for this upload
        run_result = await db.execute(
            select(ExtractionRun).where(ExtractionRun.config_version == upload_id)
        )
        runs = run_result.scalars().all()
        if not runs:
            return JSONResponse(status_code=404, content={"error": "No extraction data found"})

        run_ids = [r.id for r in runs]
        row_result = await db.execute(
            select(ExtractedRow)
            .where(ExtractedRow.extraction_run_id.in_(run_ids))
            .order_by(ExtractedRow.created_at)
        )
        rows = row_result.scalars().all()

        # Load the customer template (column order + checklist + explanations)
        template = _export_template()
        column_specs = template.get("columns") or [
            {"field": f, "label": l, "area": "", "required": ""}
            for f, l in _EXCEL_FIELDS
        ]
        explanations = template.get("explanations") or {}
        checklist_items = template.get("checklist") or []
        req_labels = template.get("requirement_labels") or {}

        wb = openpyxl.Workbook()
        _build_extracted_data_sheet(wb.active, rows, column_specs, req_labels)
        _build_checklist_sheet(wb.create_sheet("Checklist"), checklist_items)
        _build_column_explanations_sheet(
            wb.create_sheet("Column Explanations"), column_specs, explanations
        )
        _build_mandatory_fields_sheet(
            wb.create_sheet("Mandatory Fields"), column_specs, req_labels
        )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventory_{upload_id}.xlsx"},
        )

    except Exception as e:
        logger.exception("Excel export failed for %s", upload_id)
        return JSONResponse(status_code=500, content={"error": str(e)})


# ────────────────────────────────────────────────────────────────────────
# Sheet builders
# ────────────────────────────────────────────────────────────────────────


def _coerce_to_column_type(model_cls, field_name: str, raw: str):
    """Cast a string to the ORM column's Python type so typed columns
    (Integer, Numeric, Date) accept the value via setattr.

    Returns None when coercion fails OR when the column doesn't exist.
    Plain-text columns return the trimmed string unchanged.
    """
    from sqlalchemy import Integer, Numeric, Float, Boolean, Date, DateTime
    from decimal import Decimal, InvalidOperation
    from datetime import date, datetime

    if not hasattr(model_cls, field_name):
        return None
    try:
        col = getattr(model_cls, field_name).property.columns[0]
        col_type = col.type
    except (AttributeError, IndexError):
        return raw

    s = (raw or "").strip()
    if not s:
        return None

    if isinstance(col_type, Integer):
        try:
            return int(float(s))  # tolerate "1.0" → 1
        except (ValueError, TypeError):
            return None
    if isinstance(col_type, (Numeric, Float)):
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError, TypeError):
            return None
    if isinstance(col_type, Boolean):
        return s.lower() in ("yes", "true", "y", "1")
    if isinstance(col_type, Date):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    if isinstance(col_type, DateTime):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None
    return s  # String / Text / JSONB / etc.


def _values_match_for_diff(a, b) -> bool:
    """Decide whether an Excel-cell value should be treated as equal to a DB
    value for the purpose of correction-diffing.

    Handles the round-trip noise that the previous import accidentally treated
    as corrections — Decimal('57.50') exported as float 57.5, dates as
    `2019-12-31` strings vs `date(2019,12,31)` objects, blank-vs-None, etc.
    Returns True if no real change was made by the analyst.
    """
    # Treat blanks as equal to blanks
    a_blank = a is None or (isinstance(a, str) and not a.strip())
    b_blank = b is None or (isinstance(b, str) and not b.strip())
    if a_blank and b_blank:
        return True
    if a_blank != b_blank:
        return False

    # Numeric tolerance — Decimals round-trip via float lose trailing zeros
    try:
        af = float(a) if not isinstance(a, bool) else None
        bf = float(b) if not isinstance(b, bool) else None
        if af is not None and bf is not None:
            return abs(af - bf) < 0.005  # half-cent tolerance for $ amounts
    except (TypeError, ValueError):
        pass

    # Date / datetime tolerance — Excel may give strings or datetimes
    from datetime import date, datetime
    def _to_date(x):
        if isinstance(x, date) and not isinstance(x, datetime):
            return x
        if isinstance(x, datetime):
            return x.date()
        if isinstance(x, str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
                try:
                    return datetime.strptime(x.strip(), fmt).date()
                except ValueError:
                    continue
        return None
    da, db_ = _to_date(a), _to_date(b)
    if da and db_:
        return da == db_

    # Generic string compare with whitespace folding
    return str(a).strip() == str(b).strip()


def _detect_header_layout(ws, known_labels: set[str]) -> tuple[int | None, int]:
    """Find the row that holds the column-display-name headers.

    Returns (header_row, data_start_row). Walks the first 5 rows and picks
    whichever row matches the most known labels — handles both the new
    customer-template export (row 3 is headers) and the legacy single-row
    header export (row 1 is headers). If nothing matches, returns
    (None, 2) so the caller can decide.
    """
    best_row = None
    best_hits = 0
    for r in range(1, min(ws.max_row, 5) + 1):
        hits = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() in known_labels:
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_row = r
    if best_row is None or best_hits < 3:  # need at least 3 recognized columns to trust the layout
        return None, 2
    return best_row, best_row + 1


def _build_extracted_data_sheet(ws, rows, column_specs, req_labels):
    """Sheet 1 — the actual extracted data, in customer column order.

    Layout:
      row 1: Area headers (Location, Carrier Information, Service, ...)
      row 2: Required level (Required, Required if Applicable, ...)
      row 3: Column display names
      row 4+: Data rows, with confidence color fill per cell
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    ws.title = "Extracted Data"

    fills = {
        "high":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "low":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    area_fill   = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    req_fill    = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    white_bold  = Font(color="FFFFFF", bold=True)
    bold        = Font(bold=True)

    # Row 1 — area headers spanning each contiguous run of same area
    last_area = None
    span_start = 1
    for col_idx, spec in enumerate(column_specs, start=1):
        area = spec.get("area") or ""
        if area != last_area:
            if last_area is not None and col_idx > span_start:
                ws.merge_cells(start_row=1, start_column=span_start, end_row=1, end_column=col_idx - 1)
                cell = ws.cell(row=1, column=span_start)
                cell.value = last_area
                cell.fill = area_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")
            last_area = area
            span_start = col_idx
    # close the last area span
    end_col = len(column_specs)
    if end_col >= span_start:
        if end_col > span_start:
            ws.merge_cells(start_row=1, start_column=span_start, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=span_start)
        cell.value = last_area
        cell.fill = area_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")

    # Row 2 — requirement level
    for col_idx, spec in enumerate(column_specs, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = req_labels.get(spec.get("required", ""), spec.get("required", ""))
        cell.fill = req_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Row 3 — column labels
    for col_idx, spec in enumerate(column_specs, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = spec["label"]
        cell.fill = header_fill
        cell.font = bold

    ws.row_dimensions[2].height = 32

    # Row 4+ — data rows, colored per-cell by confidence
    text_cols = []
    for col_idx, spec in enumerate(column_specs, start=1):
        if spec["field"] in ("btn", "phone_number", "carrier_account_number",
                             "sub_account_number_1", "sub_account_number_2",
                             "zip", "z_zip", "carrier_circuit_number"):
            text_cols.append(col_idx)

    data_start_row = 4
    for r_idx, row in enumerate(rows, start=data_start_row):
        confidence_map = row.field_confidence or {}
        for col_idx, spec in enumerate(column_specs, start=1):
            val = getattr(row, spec["field"], None)
            if val is None:
                continue
            # compliance_flags is a JSONB list of {check, severity, message,
            # details} — render as one human-readable string per row so the
            # analyst sees why a row was flagged without expanding JSON.
            if spec["field"] == "compliance_flags":
                # Empty list (the JSONB column default) → skip the cell.
                if not isinstance(val, list) or not val:
                    continue
                lines = []
                for f in val:
                    if not isinstance(f, dict):
                        continue
                    check = f.get("check") or ""
                    sev   = f.get("severity") or ""
                    msg   = f.get("message") or ""
                    lines.append(f"[{sev}] {check}: {msg}")
                if not lines:
                    continue
                val = "\n".join(lines)

            # Generic safety net: any other list/dict slipping through (from
            # JSONB columns we add later) gets stringified rather than
            # crashing openpyxl. Pure scalars pass through unchanged.
            elif isinstance(val, (list, dict)):
                if not val:
                    continue
                import json as _json
                val = _json.dumps(val, default=str)

            cell = ws.cell(row=r_idx, column=col_idx)
            cell.value = val
            field_conf = confidence_map.get(spec["field"])
            if field_conf in fills:
                cell.fill = fills[field_conf]
            if col_idx in text_cols:
                cell.number_format = "@"
            if spec["field"] == "compliance_flags":
                # Show severity-coded fill so flagged rows stand out at a glance
                from openpyxl.styles import Alignment as _Align
                cell.alignment = _Align(wrap_text=True, vertical="top")
                if "[error]" in val.lower():
                    cell.fill = fills.get("low") or cell.fill
                elif "[warning]" in val.lower():
                    cell.fill = fills.get("medium") or cell.fill

    # Reasonable column widths
    for col_idx, spec in enumerate(column_specs, start=1):
        col_letter = ws.cell(row=3, column=col_idx).column_letter
        # Wider for known long columns
        long_cols = {"notes", "auto_renewal_notes", "files_used", "additional_circuit_ids",
                     "component_or_feature_name", "billing_name", "service_address_1",
                     "z_address_1", "contract_file_name", "z_location_name"}
        if spec["field"] == "compliance_flags":
            ws.column_dimensions[col_letter].width = 50
        elif spec["field"] in long_cols:
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 18

    ws.freeze_panes = "A4"


def _build_checklist_sheet(ws, checklist_items):
    """Sheet 2 — 30 QA items + Agent / QA Yes-No columns blank for analyst to fill."""
    from openpyxl.styles import PatternFill, Font, Alignment

    ws.title = "Checklist"
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)

    headers = ["Checklist", "Agent — Yes/No", "QA — Yes/No"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")

    for r_idx, item in enumerate(checklist_items, start=2):
        ws.cell(row=r_idx, column=1).value = item
        ws.cell(row=r_idx, column=1).alignment = Alignment(wrap_text=True)
        # Agent + QA columns left blank intentionally; analyst fills in.

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[1].height = 22


def _build_column_explanations_sheet(ws, column_specs, explanations):
    """Sheet 3 — column name + explanation, exactly as the customer ships it."""
    from openpyxl.styles import PatternFill, Font, Alignment

    ws.title = "Column Explanations"
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)

    headers = ["Column Name", "Explanation"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")

    for r_idx, spec in enumerate(column_specs, start=2):
        ws.cell(row=r_idx, column=1).value = spec["label"]
        explanation = explanations.get(spec["field"]) or ""
        ws.cell(row=r_idx, column=2).value = explanation
        ws.cell(row=r_idx, column=2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80


def _build_mandatory_fields_sheet(ws, column_specs, req_labels):
    """Sheet 4 — requirement matrix grouped by area, mirroring the customer's
    "Mandatory Field details" tab structure."""
    from openpyxl.styles import PatternFill, Font, Alignment

    ws.title = "Mandatory Fields"
    area_fill   = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    white_bold  = Font(color="FFFFFF", bold=True)
    bold        = Font(bold=True)

    headers = ["Area", "Column", "Requirement Level"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = area_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")

    last_area = None
    for r_idx, spec in enumerate(column_specs, start=2):
        area = spec.get("area") or ""
        # Only repeat area on first row of each area group
        if area != last_area:
            ws.cell(row=r_idx, column=1).value = area
            ws.cell(row=r_idx, column=1).font = bold
            last_area = area
        ws.cell(row=r_idx, column=2).value = spec["label"]
        ws.cell(row=r_idx, column=3).value = req_labels.get(
            spec.get("required", ""), spec.get("required", "")
        )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 56


@router.post("/corrections/import")
async def import_corrections(
    upload_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload corrected Excel, diff against extraction, store corrections.

    Reads the workbook label-driven (not position-driven) so it stays in sync
    with the customer-template layout produced by `export_excel`:

      Sheet 1 "Extracted Data" — area headers in row 1, required-level in row 2,
      column display names in row 3, data starting at row 4.

    The previous implementation hard-coded a 59-column position list and
    `min_row=2`; the new export ships 61 columns + 3 header rows so positional
    parsing was misaligning every row by 2 and silently writing wrong fields
    into the corrections table. Matt's `image010` error came from here.

    Diffing strategy:
      - Map Excel column LABELS (row 3) back to ORM field names via the
        export_template.yaml mapping. Any label we don't recognize is skipped
        (forward-compatible if the customer adds a column to their template).
      - Skip the rendered "Compliance Audit Flags" column on import — it's
        platform-computed, not analyst input.
      - Per-cell diff: when the Excel cell differs from the current DB value,
        record a Correction (with embedding generation deferred to /api/review),
        then write the corrected value onto the row.
    """
    try:
        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))

        # Pick the data sheet — prefer "Extracted Data" if present (new export),
        # fall back to the active sheet (older single-sheet export).
        if "Extracted Data" in wb.sheetnames:
            ws = wb["Extracted Data"]
        else:
            ws = wb.active

        # Build label → ORM field map from the export template (single source
        # of truth). Falls back to the legacy _EXCEL_FIELDS list if the YAML
        # is missing — keeps old workbooks importable.
        template = _export_template()
        column_specs = template.get("columns") or [
            {"field": f, "label": l} for f, l in _EXCEL_FIELDS
        ]
        label_to_field = {spec["label"]: spec["field"] for spec in column_specs}
        # Ignore platform-computed columns the analyst can't meaningfully edit.
        IGNORE_LABELS = {"Compliance Audit Flags"}

        # Detect which row holds the column-name headers + which row data starts.
        # New export: row 1 = area, row 2 = required-level, row 3 = labels, row 4+ = data
        # Legacy export: row 1 = labels, row 2+ = data
        header_row, data_start_row = _detect_header_layout(ws, set(label_to_field.keys()))
        if header_row is None:
            return JSONResponse(status_code=400, content={
                "error": "Could not locate column headers in the uploaded workbook. "
                         "Use the export downloaded from this platform as the import template."
            })

        # Build column-index → ORM field name map for THIS workbook
        col_to_field: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            label = ws.cell(row=header_row, column=c).value
            if not label:
                continue
            label = str(label).strip()
            if label in IGNORE_LABELS:
                continue
            if label in label_to_field:
                col_to_field[c] = label_to_field[label]
        if not col_to_field:
            return JSONResponse(status_code=400, content={
                "error": "No recognized column headers in the uploaded workbook. "
                         "Re-download the inventory from this platform and re-upload."
            })

        # Find extraction run(s)
        run_result = await db.execute(
            select(ExtractionRun).where(ExtractionRun.config_version == upload_id)
        )
        runs = run_result.scalars().all()
        if not runs:
            return JSONResponse(status_code=404, content={"error": "No extraction data found for this upload"})

        run_ids = [r.id for r in runs]
        row_result = await db.execute(
            select(ExtractedRow)
            .where(ExtractedRow.extraction_run_id.in_(run_ids))
            .order_by(ExtractedRow.created_at)
        )
        db_rows = row_result.scalars().all()
        if not db_rows:
            return JSONResponse(status_code=404, content={"error": "No rows in this upload to compare against"})

        # Diff each Excel row against the matching DB row
        corrections_created = 0
        rows_compared = 0
        for idx, excel_row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True)):
            if idx >= len(db_rows):
                break
            db_row = db_rows[idx]
            rows_compared += 1
            row_changed = False

            for col_idx, field_name in col_to_field.items():
                excel_val = excel_row[col_idx - 1] if col_idx - 1 < len(excel_row) else None
                db_val = getattr(db_row, field_name, None)

                if _values_match_for_diff(excel_val, db_val):
                    continue

                excel_str = str(excel_val).strip() if excel_val not in (None, "") else None
                db_str = str(db_val).strip() if db_val not in (None, "") else None
                # Skip blank-Excel-cell-against-populated-DB-value: an analyst
                # who DELETES a value should use the UI; an empty cell on
                # import is more often "didn't touch it" than "wanted to clear".
                if excel_str is None:
                    continue
                # Coerce to the ORM column's Python type so typed columns
                # (Integer, Numeric, Date) don't blow up when given a string.
                # Falls back to the raw string for plain-text columns.
                coerced = _coerce_to_column_type(ExtractedRow, field_name, excel_str)
                if coerced is None and excel_str:
                    # Couldn't coerce — log but don't crash the import; record the
                    # correction with the raw text so the analyst's intent is captured.
                    coerced = excel_str
                db.add(Correction(
                    extracted_row_id=db_row.id,
                    extraction_run_id=db_row.extraction_run_id,
                    field_name=field_name,
                    extracted_value=db_str,
                    corrected_value=excel_str,
                    correction_type="excel_import",
                    carrier=db_row.carrier,
                ))
                try:
                    setattr(db_row, field_name, coerced)
                except (TypeError, ValueError) as e:
                    logger.warning(
                        "import_corrections: skipped writing %s=%r on row %s: %s",
                        field_name, coerced, db_row.id, e,
                    )
                    continue
                corrections_created += 1
                row_changed = True

            if row_changed:
                db_row.review_status = "corrected"

        await db.commit()
        return {
            "upload_id": upload_id,
            "rows_compared": rows_compared,
            "corrections_created": corrections_created,
            "columns_recognized": len(col_to_field),
            "header_row": header_row,
            "data_start_row": data_start_row,
        }

    except Exception as e:
        logger.error(f"Corrections import failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})
